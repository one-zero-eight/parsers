"""
This file should be synced between:
https://github.com/one-zero-eight/parsers/blob/main/src/core_courses/parser.py
https://github.com/one-zero-eight/schedule-builder-backend/blob/main/src/core_courses/parser.py
"""

import datetime
import io
import re
from collections import defaultdict
from collections.abc import Generator
from itertools import pairwise

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
from pandas.core.frame import DataFrame
from pydantic import BaseModel, ConfigDict, Field

from src.logging_ import logger

from ..utils import WEEKDAYS, prettify_string, sanitize_sheet_name


class CoreCourseCell(BaseModel):
    model_config = ConfigDict(coerce_numbers_to_str=True)

    value: tuple[str, str | None, str | None] = Field(..., min_length=3, max_length=3)
    "Cell values: [subject, teacher (optional), location and modifiers (optional)]"
    spreadsheet_id: str
    "Spreadsheet ID"
    google_sheet_gid: str
    "Sheet GID"
    google_sheet_name: str
    "Sheet name"
    a1: str | None
    "A1 coordinates of left-upper cell"

    def __repr__(self):
        return "\n".join(map(str, self.value))


class CoreCoursesParser:

    def __init__(self):
        self.last_dfs_merged_ranges: dict[str, list[tuple[int, int, int, int]]] | None = None

    def pipeline(
        self,
        xlsx_file: io.BytesIO,
        original_target_sheet_names: list[str],
        sheet_gids: dict[str, str],
        spreadsheet_id: str,
    ) -> Generator[list[DataFrame], None, None]:
        """
        Run pipeline and generate lists of GroupBy with CoreCourseCell(value=[subject, teacher, location], a1=excel_range) by sheet.

        ### Usage:

        ```python
        pipeline_result = parser.pipeline(xlsx, sheet_names, sheet_gids, spreadsheet_id)

        def use(processed_column: pd.Series, some_variable: str):
            \"\"\"
            :param processed_column: series with processed cells (CoreCourseCell),
                multiindex with (weekday, timeslot) and (course, group) as name
            \"\"\"
            (course, group) = processed_column.name
            course: str
            group: str

            for (weekday, timeslot), cell in processed_column.items():
                cell: CoreCourseCell | None
                if cell is None:
                    continue
                weekday: str
                timeslot: tuple[datetime.time, datetime.time]
                yield cell # Do what you want with cell

        all_cells = []
        some_variable = "hello there"

        for sheet_name, grouped_dfs_with_cells_list in zip(sheet_names, pipeline_result):
            for grouped_dfs_with_cells in grouped_dfs_with_cells_list:
                series_with_generators = grouped_dfs_with_cells.apply(use, some_variable=some_variable)
                for generator in series_with_generators:
                    generator: Generator[CoreCourseCell, None, None]
                    all_cells.extend(generator)

        ```
        """

        sanitized_sheet_names = [
            sanitize_sheet_name(target_sheet_name) for target_sheet_name in original_target_sheet_names
        ]

        sanitized_sheet_name_x_google_sheet_name = {
            sanitize_sheet_name(sheet_name): sheet_name for sheet_name in sheet_gids.keys()
        }

        dfs, self.last_dfs_merged_ranges = self.get_clear_dataframes_from_xlsx(
            xlsx_file=xlsx_file, target_sheet_names=sanitized_sheet_names
        )

        for target_sheet_name in sanitized_sheet_names:
            # find dataframe from dfs
            if target_sheet_name not in dfs:
                logger.warning(f"Sheet {target_sheet_name} not found in xlsx file")
                continue
            sheet_df = dfs[target_sheet_name]
            google_sheet_name = sanitized_sheet_name_x_google_sheet_name.get(target_sheet_name)
            google_sheet_gid = sheet_gids.get(google_sheet_name) if google_sheet_name else None

            time_columns_index = self.get_time_columns(sheet_df)
            logger.info(f"Sheet Time columns: {[get_column_letter(col + 1) for col in time_columns_index]}")
            rightmost_column_index = self.get_rightmost_column_index(xlsx_file, target_sheet_name, time_columns_index)
            logger.info(f"Rightmost column index: {get_column_letter(rightmost_column_index + 1)}")

            by_courses = self.split_df_by_courses(sheet_df, time_columns_index)
            grouped_dfs_with_cells_lst = []
            for course_df in by_courses:
                # ---- Set course and group as header; weekday and timeslot as index ----
                self.set_course_and_group_as_header(course_df)
                self.set_weekday_and_time_as_index(course_df)
                # ---- Convert it to GroupBy with CoreCourseCell(value=[subject, teacher, location], a1=excel_range) ----
                grouped_dfs_with_cells = (
                    course_df
                    # ---- Group by weekday and time ----
                    .groupby(level=[0, 1], sort=False)
                    .agg(list)
                    # ---- Convert each cell to CoreCourseCell ----
                    .map(
                        self.factory_core_course_cell,
                        spreadsheet_id=spreadsheet_id,
                        google_sheet_name=google_sheet_name,
                        google_sheet_gid=google_sheet_gid,
                    )
                )
                assert isinstance(grouped_dfs_with_cells, DataFrame)
                grouped_dfs_with_cells_lst.append(grouped_dfs_with_cells)
            yield grouped_dfs_with_cells_lst

    def get_clear_dataframes_from_xlsx(
        self, xlsx_file: io.BytesIO, target_sheet_names: list[str]
    ) -> tuple[dict[str, pd.DataFrame], dict]:
        """
        Get data from xlsx file and return it as a DataFrame with merged
        cells and empty cells in the course row filled by left value.
        Also adds excel range to each 'subject' cell (first of three cells),
        so will be `Analytical Geometry and Linear Algebra I (lab)$D10`

        :return: mapping of sheet name to clear dataframe
        :rtype: dict[str, pd.DataFrame]
        """
        # ---- Read xlsx file into dataframes ----
        dfs = pd.read_excel(xlsx_file, engine="openpyxl", sheet_name=None, header=None)
        # ---- Clean up dataframes ----
        merged_ranges: dict[str, list[tuple[int, int, int, int]]] = defaultdict(list)
        for target_sheet_name in target_sheet_names:
            df = dfs[target_sheet_name]
            # ---- Select range ----
            (min_row, min_col, max_row, max_col) = self.auto_detect_range(df, xlsx_file, target_sheet_name)
            df = df.iloc[min_row : max_row + 1, min_col : max_col + 1]
            # ---- Fill merged cells with values ----
            merged_ranges[target_sheet_name] = self.merge_cells(df, xlsx_file, target_sheet_name)
            # ---- Add excel range to each 'subject' cell (first of three cells) ----
            self.assign_excel_row_and_column_to_subject(df)
            # ---- Fill empty cells ----
            df = df.replace(r"^\s*$", np.nan, regex=True)
            # ---- Strip, translate and remove trailing spaces ----
            df = df.map(prettify_string)
            # ---- Update dataframe ----
            dfs[target_sheet_name] = df

        return dfs, merged_ranges

    def auto_detect_range(
        self, sheet_df: pd.DataFrame, xlsx_file: io.BytesIO, sheet_name: str
    ) -> tuple[int, int, int, int]:
        """
        :return: tuple of (min_row, min_col, max_row, max_col)
        """
        time_columns_index = self.get_time_columns(sheet_df)
        logger.info(f"Time columns: {[get_column_letter(col + 1) for col in time_columns_index]}")
        # ---- Get rightmost column index ----
        rightmost_column_index = self.get_rightmost_column_index(xlsx_file, sheet_name, time_columns_index)
        logger.info(f"Rightmost column index: {get_column_letter(rightmost_column_index + 1)}")
        last_row_index = self.get_last_row_index(xlsx_file, sheet_name)
        target_range = f"A1:{get_column_letter(rightmost_column_index + 1)}{last_row_index}"
        logger.info(f"Target range: {target_range}")
        return (0, 0, last_row_index, rightmost_column_index)

    def get_time_columns(self, sheet_df: pd.DataFrame) -> list[int]:
        # find columns where presents "MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY"
        time_columns = []
        for column in sheet_df.columns:
            values_in_column = sheet_df[column].values
            if all(weekday in values_in_column for weekday in WEEKDAYS[:-1]):
                time_columns.append(column)
        return time_columns

    def get_rightmost_column_index(self, xlsx_file: io.BytesIO, sheet_name: str, time_columns: list[int]) -> int:
        # Column after time columns that has no borders formatting

        wb = openpyxl.load_workbook(xlsx_file)
        sheet = wb[sheet_name]
        last_time_column = time_columns[-1]

        next_column = last_time_column + 1
        cell = sheet.cell(row=1, column=next_column + 1)

        # Check if cell has no borders
        has_no_border = (
            (cell.border is None or cell.border.right is None or cell.border.right.style is None)
            and (cell.border is None or cell.border.top is None or cell.border.top.style is None)
            and (cell.border is None or cell.border.bottom is None or cell.border.bottom.style is None)
        )

        if has_no_border:
            return next_column - 1
        else:
            # Continue searching for column with no border
            for col in range(next_column + 1, sheet.max_column + 1):
                cell = sheet.cell(row=1, column=col + 1)
                if (
                    (cell.border is None or cell.border.right is None or cell.border.right.style is None)
                    and (cell.border is None or cell.border.top is None or cell.border.top.style is None)
                    and (cell.border is None or cell.border.bottom is None or cell.border.bottom.style is None)
                ):
                    return col - 1
            return next_column  # fallback

    def get_last_row_index(self, xlsx_file: io.BytesIO, sheet_name: str) -> int:
        wb = openpyxl.load_workbook(xlsx_file)
        sheet = wb[sheet_name]
        return sheet.max_row

    def assign_excel_row_and_column_to_subject(self, df: pd.DataFrame) -> None:
        def check_value_is_time(string_to_check: str) -> bool:
            return bool(re.match(r"^\d{1,2}:\d{2}\s*-\s*\d{1,2}:\d{2}$", string_to_check))

        used_cells: set[tuple[int, int]] = set()
        for i in range(3, len(df.values)):
            for j in range(1, len(df.values[i])):
                if (i, j) in used_cells:
                    continue

                v = df.iloc[i, j]
                if isinstance(v, str):
                    v = v.strip()

                if not v or pd.isna(v) or v in WEEKDAYS or check_value_is_time(v):
                    continue

                excel_coords = f"{get_column_letter(j + 1)}{i + 1}"
                df.iloc[i, j] = f"{df.iloc[i, j]}${excel_coords}"
                for x in range(i, i + 3):
                    used_cells.add((x, j))

    def merge_cells(
        self, df: pd.DataFrame, xlsx: io.BytesIO, target_sheet_name: str
    ) -> list[tuple[int, int, int, int]]:
        """
        Merge cells in dataframe

        :param df: Dataframe to process
        :param xlsx: xlsx file with data
        :param target_sheet_name: sheet to process
        :return: list of merged ranges: (min_row, min_col, max_row, max_col)
        """
        xlsx.seek(0)
        ws = openpyxl.load_workbook(xlsx)
        sheet = ws[target_sheet_name]

        merged_ranges = []
        nrows, ncols = df.shape

        def clamp_rows(n: int) -> int:
            return max(min(n, nrows - 1), 0)

        def clamp_cols(n: int) -> int:
            return max(min(n, ncols - 1), 0)

        for merged_range in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            min_col = clamp_cols(min_col - 1)
            min_row = clamp_rows(min_row - 1)
            max_col = clamp_cols(max_col - 1)
            max_row = clamp_rows(max_row - 1)

            value = df.iloc[min_row, min_col]
            df.iloc[min_row : max_row + 1, min_col : max_col + 1] = value
            merged_ranges.append((min_row, min_col, max_row, max_col))

        return merged_ranges

    def set_weekday_and_time_as_index(self, df: pd.DataFrame, column: int = 0) -> None:
        """
        Set time column as index and process it to datetime format

        :param df: dataframe to process
        :type df: pd.DataFrame
        :param column: column to set as index, defaults to 0
        :type column: int, optional
        """

        # get column view and iterate over it
        df_column = df.iloc[:, column]
        df_column: pd.Series
        # drop column
        df.drop(df.columns[column], axis=1, inplace=True)
        # fill nan values with previous value
        df_column.ffill(inplace=True)

        # ----- Process weekday ------ #
        # get indexes of weekdays
        weekdays_indexes = [i for i, cell in enumerate(df_column.values) if cell in WEEKDAYS]

        # create index mapping for weekdays [None, None, "MONDAY", "MONDAY", ...]
        index_mapping = pd.Series(index=df_column.index, dtype=object)
        last_index = len(df_column)
        for start, end in pairwise(weekdays_indexes + [last_index]):
            index_mapping.iloc[start] = "delete"
            index_mapping.iloc[start + 1 : end] = df_column[start]

        # ----- Process time ------ #
        matched = df_column[df_column.str.match(r"\d{1,2}:\d{2}-\d{1,2}:\d{2}")]

        for i, cell in matched.items():
            # "9:00-10:30" -> datetime.time(9, 0), datetime.time(10, 30)
            start, end = cell.split("-")
            df_column.loc[i] = (
                datetime.datetime.strptime(start, "%H:%M").time(),
                datetime.datetime.strptime(end, "%H:%M").time(),
            )

        # create multiindex from index mapping and time column
        multiindex = pd.MultiIndex.from_arrays([index_mapping, df_column], names=["weekday", "time"])
        # set multiindex as index
        df.set_index(multiindex, inplace=True)
        # drop rows with weekday
        df.drop("delete", inplace=True, level=0)

    def set_course_and_group_as_header(self, df: pd.DataFrame, rows: tuple = (0, 1)) -> None:
        """
        Set course and group as header

        :param df: dataframe to process
        :type df: pd.DataFrame
        :param rows: row to set as columns, defaults to (0, 1)
        :type rows: tuple, optional
        """
        # --- Set course and group as header ---
        # get rows with course and group
        df_header = df.iloc[rows[0] : rows[1] + 1]
        # drop rows with course and group
        df.drop(list(rows), inplace=True)
        df.reset_index(drop=True, inplace=True)
        # fill nan values with previous value
        with pd.option_context("future.no_silent_downcasting", True):
            df_header = df_header.ffill(axis=1)
        multiindex = pd.MultiIndex.from_arrays(df_header.values, names=["course", "group"])
        df.columns = multiindex

    def factory_core_course_cell(
        self,
        values: list[str | None],
        google_sheet_name: str,
        google_sheet_gid: str,
        spreadsheet_id: str,
    ) -> CoreCourseCell | None:
        if all(pd.isna(y) for y in values):
            return None
        if len(values) == 3:
            values = [None if (bool(pd.isna(x))) else x for x in values]
        elif len(values) == 1:
            values = [None if (bool(pd.isna(values[0]))) else values[0]] + [None] * 2
        else:
            raise ValueError(f"Length of value must be 3 or 1, got {values}")

        assert values[0] is not None, f"Subject must not be None, got {values}"
        assert len(values) == 3, f"Length of value must be 3, got {values}"

        a1 = None
        for i, v in enumerate(values):
            if v is not None and isinstance(v, str) and "$" in v:
                values[i], a1 = v.rsplit("$", maxsplit=1)
        return CoreCourseCell(
            value=tuple(values),
            spreadsheet_id=spreadsheet_id,
            google_sheet_name=google_sheet_name,
            google_sheet_gid=google_sheet_gid,
            a1=a1,
        )

    def split_df_by_courses(self, df: pd.DataFrame, time_columns: list[int]) -> list[pd.DataFrame]:
        """
        Split dataframe by "Week *" rows

        :param time_columns: list of columns(pd) with time and weekday
        :param df: dataframe to split
        :type df: pd.DataFrame
        :return: list of dataframes
        :rtype: list[pd.DataFrame]
        """

        logger.debug("Parsing dataframe to separation by course|groups...")
        logger.info("Get indexes of time columns...")

        time_columns_indexes = time_columns
        logger.info(f"Time columns indexes: {time_columns_indexes}")

        # split dataframe by found columns
        _, max_y = df.shape
        split_indexes = time_columns_indexes + [max_y]

        split_dfs = []

        for i, (start, end) in enumerate(pairwise(split_indexes)):
            logger.info(f"Splitting dataframe by columns {start}:{end}")
            split_df = df.iloc[:, start:end].copy()
            split_dfs.append(split_df)
        return split_dfs
