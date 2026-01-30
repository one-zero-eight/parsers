"""
This file should be synced between:
https://github.com/one-zero-eight/parsers/blob/main/src/electives/parser.py
https://github.com/one-zero-eight/schedule-builder-backend/blob/main/src/parsers/electives/parser.py
"""

import datetime
import io
import re
import warnings
from collections.abc import Generator
from itertools import groupby, pairwise

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from src.logging_ import logger

from ..utils import prettify_string, sanitize_sheet_name
from .cell_to_event import ElectiveEvent
from .config import Elective

BRACKETS_PATTERN = re.compile(r"\((.*?)\)")


class ElectiveCell(BaseModel):
    value: list[str]
    "Original cell value"
    a1: str | None = None
    "A1 coordinates of the cell"

    def __repr__(self):
        return "\n".join(self.value)


class Separation(BaseModel):
    elective: Elective
    events: list[ElectiveEvent]


class ElectiveParser:
    """
    Elective parser class
    """

    def pipeline(
        self,
        xlsx_file: io.BytesIO,
        original_target_sheet_names: list[str],
        electives: list[Elective],
    ) -> Generator[list[Separation], None, None]:
        sanitized_target_sheet_names = [
            sanitize_sheet_name(target_sheet_name) for target_sheet_name in original_target_sheet_names
        ]
        dfs = self.get_clear_dataframes_from_xlsx(xlsx_file, sanitized_target_sheet_names)

        for target_sheet_name, original_target_sheet_name in zip(
            sanitized_target_sheet_names, original_target_sheet_names
        ):
            # find dataframe from dfs
            if target_sheet_name not in dfs:
                logger.warning(f"Sheet {target_sheet_name} not found in xlsx file")
                continue
            sheet_df = dfs[target_sheet_name]

            by_weeks = self.split_df_by_weeks(sheet_df)
            index = {}
            for sheet_df in by_weeks:
                index.update(sheet_df.index.tolist())
            big_df = pd.DataFrame(index=index)
            big_df = pd.concat([big_df, *by_weeks], axis=1)
            big_df.dropna(axis=1, how="all", inplace=True)
            big_df.dropna(axis=0, how="all", inplace=True)
            all_events = list(self.parse_df(big_df, electives, original_target_sheet_name))
            converted = self.events_to_separation_by_elective(all_events)
            yield converted

    def get_clear_dataframes_from_xlsx(
        self, xlsx_file: io.BytesIO, target_sheet_names: list[str]
    ) -> dict[str, pd.DataFrame]:
        """
        Get data from xlsx file and return it as a DataFrame with merged
        cells and empty cells in the course row filled by left value.

        :param xlsx_file: xlsx file with data
        :type xlsx_file: io.BytesIO
        :param target_sheet_names: list of target sheet names to get data from
        :type target_sheet_names: list[str]

        :return: dataframes with merged cells and empty cells filled
        :rtype: dict[str, pd.DataFrame]
        """
        # ------- Read xlsx file into dataframes -------
        dfs = pd.read_excel(xlsx_file, engine="openpyxl", sheet_name=None, header=None)

        # ------- Clean up dataframes -------
        for target_sheet_name in target_sheet_names:
            df = dfs[target_sheet_name]
            # -------- Select range --------
            (min_row, min_col, max_row, max_col) = self.auto_detect_range(df, xlsx_file, target_sheet_name)

            # -------- Add Excel coordinates to cell values --------
            df = df.iloc[min_row : max_row + 1, min_col : max_col + 1]
            self.assign_excel_row_and_column_to_cells(df, min_row, min_col)
            # -------- Set time column as index --------
            df = self.set_time_column_as_index(df)
            # -------- Strip all values --------
            df = df.map(lambda x: x.strip() if isinstance(x, str) else x)
            # -------- Fill empty cells --------
            df = df.replace(r"^\s*$", np.nan, regex=True)
            # -------- Exclude nan rows --------
            df = df.dropna(how="all")
            # -------- Strip, translate and remove trailing spaces --------
            df = df.map(prettify_string)
            # -------- Update dataframe --------
            dfs[target_sheet_name] = df

        return dfs

    def events_to_separation_by_elective(self, events: list[ElectiveEvent]) -> list[Separation]:
        """
        Convert list of events to dict with separation by Elective.

        :param events: list of events to convert
        :type events: list[ElectiveEvent]
        :return: separations by Elective
        :rtype: list[Separation]
        """
        output: dict[str, Separation] = dict()

        # # by groups of elective
        # for (elective, group), _events in groupby(events, lambda e: (e.elective, e.group)):
        #     elective: Elective
        #     if group is None:
        #         continue  # cal = output[elective.alias]
        #     else:
        #         cal = output[f"{elective.alias}-{group}"]
        #     cal: list[ElectiveEvent]
        #     cal.extend(_events)

        # only by Elective
        for elective, _events in groupby(events, lambda e: e.elective):
            elective: Elective
            if elective.alias not in output:
                output[elective.alias] = Separation(
                    elective=elective,
                    events=list(_events),
                )
            else:
                output[elective.alias].events.extend(_events)

        return list(output.values())

    def auto_detect_range(
        self, sheet_df: pd.DataFrame, xlsx_file: io.BytesIO, sheet_name: str
    ) -> tuple[int, int, int, int]:
        """
        :return: tuple of (min_row, min_col, max_row, max_col)
        """

        weekdays = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY", "SUNDAY"]

        # find all columns named as weekday by checking the first row
        first_row = sheet_df.iloc[0]
        weekday_columns_index = [
            i for i, value in enumerate(first_row) if isinstance(value, str) and value.upper().strip() in weekdays
        ]
        assert len(weekday_columns_index) == len(weekdays), "Weekday columns not found"
        rightmost_column_index = max(weekday_columns_index)
        leftmost_column_index = min(weekday_columns_index) - 1
        logger.info(f"Rightmost column index: {get_column_letter(rightmost_column_index + 1)}")
        last_row_index = self.get_last_row_index(xlsx_file, sheet_name)
        target_range = f"{get_column_letter(leftmost_column_index + 1)}1:{get_column_letter(rightmost_column_index + 1)}{last_row_index}"
        logger.info(f"Target range: {target_range}")
        return (0, leftmost_column_index, last_row_index, rightmost_column_index)

    def get_last_row_index(self, xlsx_file: io.BytesIO, sheet_name: str) -> int:
        wb = openpyxl.load_workbook(xlsx_file)
        sheet = wb[sheet_name]
        return sheet.max_row

    def set_time_column_as_index(self, df: pd.DataFrame, column: int = 0) -> pd.DataFrame:
        """
        Set time column as index and process it to datetime format

        :param df: dataframe to process
        :type df: pd.DataFrame
        :param column: column to set as index, defaults to 0
        :type column: int, optional
        """

        # "9:00-10:30" -> datetime.time(9, 0), datetime.time(10, 30)
        def process_time_cell(cell: str) -> tuple[datetime.time, datetime.time] | str:
            if "$" in cell:
                cell, a1 = cell.rsplit("$", maxsplit=1)
                cell = cell.strip()
            if re.match(r"\d{1,2}:\d{2}-\d{1,2}:\d{2}", cell):
                start, end = cell.split("-")
                return (
                    datetime.datetime.strptime(start, "%H:%M").time(),
                    datetime.datetime.strptime(end, "%H:%M").time(),
                )
            else:
                return cell

        df[column] = df[column].apply(lambda x: process_time_cell(x) if isinstance(x, str) else x)
        df.set_index(column, inplace=True)
        df.rename_axis(index="time", inplace=True)
        return df

    def set_date_row_as_header(self, df: pd.DataFrame, row: int = 0) -> pd.DataFrame:
        """
        Set date row as columns and process it to datetime format

        :param df: dataframe to process
        :type df: pd.DataFrame
        :param row: row to set as columns, defaults to 0
        :type row: int, optional
        """

        # "June 7" -> datetime.date(current_year, 6, 7)
        def process_date_cell(cell: str) -> datetime.date | str:
            if "$" in cell:
                cell, a1 = cell.rsplit("$", maxsplit=1)
                cell = cell.strip()
            if re.match(r"\w+ \d+", cell):
                dtime = datetime.datetime.strptime(cell, "%B %d")
                dtime = dtime.replace(year=datetime.date.today().year)
                return dtime.date()
            else:
                return cell

        index = df.index[row]
        df.loc[index] = df.loc[index].apply(lambda x: process_date_cell(x) if isinstance(x, str) else x)
        df.columns = df.loc[index]
        df.drop(index, inplace=True)
        df.rename_axis(columns="date", inplace=True)
        return df

    def assign_excel_row_and_column_to_cells(self, df: pd.DataFrame, min_row: int, min_col: int) -> None:
        """
        Add Excel coordinates to cell values in the format: "value$A1"

        :param df: dataframe to process
        :type df: pd.DataFrame
        :param min_row: minimum row index in original sheet (0-indexed)
        :type min_row: int
        :param min_col: minimum column index in original sheet (0-indexed)
        :type min_col: int
        """
        for i in range(len(df.values)):
            for j in range(len(df.values[i])):
                v = df.iloc[i, j]
                if pd.isna(v) or not isinstance(v, str) or not v.strip():
                    continue

                # Calculate Excel coordinates (1-indexed)
                excel_row = min_row + i + 1
                excel_col = min_col + j + 1
                excel_coords = f"{get_column_letter(excel_col)}{excel_row}"

                # Add coordinates to cell value if not already present
                if "$" not in v:
                    df.iloc[i, j] = f"{v}${excel_coords}"

    def split_df_by_weeks(self, df: pd.DataFrame) -> list[pd.DataFrame]:
        """
        Split dataframe by "Week *" rows

        :param df: dataframe to split
        :type df: pd.DataFrame
        :return: list of dataframes
        :rtype: list[pd.DataFrame]
        """

        logger.debug("Parsing dataframe to separation by days|groups...")
        logger.debug("Get 'week' indexes...")
        # find indexes of row with "Week *"
        where = df.index.str.contains(r"Week \d", na=False)
        week_indexes = df.index[where]
        week_locations = df.index.get_indexer_for(week_indexes).tolist()
        logger.debug(f"> Found {len(week_locations)} weeks")

        max_x, _ = df.shape
        week_locations += [max_x]  # add last index
        # split dataframe by week indexes
        dfs = []
        for start, end in pairwise(week_locations):
            week = df.index[start]
            logger.debug(f"Processing week: {week}... From ({start}) to ({end})")
            week_df: pd.DataFrame = df.iloc[start:end].copy()
            # ----- Set date row as header -----
            week_df = self.set_date_row_as_header(week_df)
            dfs.append(week_df)
        return dfs

    def parse_df(
        self, df: pd.DataFrame, electives: list[Elective], sheet_name: str
    ) -> Generator[ElectiveEvent, None, None]:
        """
        Parse dataframe with schedule

        :param df: dataframe with schedule
        :type df: pd.DataFrame
        :param electives: list of electives
        :type electives: list[Elective]
        :param sheet_name: name of the sheet being parsed
        :type sheet_name: str
        :return: parsed events
        """
        from .cell_to_event import convert_cell_to_events

        _elective_short_names = [e.short_name for e in electives]
        _elective_line_pattern = re.compile(r"(?P<elective_short_name>" + "|".join(_elective_short_names) + r")")

        def process_line(line: str) -> ElectiveCell | None:
            """
            Process line of the dataframe

            :param line: line to process
            :type line: str
            :return: ElectiveCell or nothing
            :rtype: ElectiveCell | None
            """
            if pd.isna(line):
                return None
            line = line.strip()

            # Extract a1 coordinates if present (format: "value$A1")
            a1 = None
            if isinstance(line, str) and "$" in line:
                line, a1 = line.rsplit("$", maxsplit=1)
                line = line.strip()

            # find all matches in the string and split by them
            matches = _elective_line_pattern.finditer(line)
            # get substrings
            breaks = [m.start() for m in matches]
            if not breaks:
                warnings.warn(
                    f"No matches found in line: {line}, most probably incorrect or missing elective short_name in config"
                )
                return ElectiveCell(value=[line], a1=a1)

            substrings = [line[i:j] for i, j in zip(breaks, breaks[1:] + [None])]
            if not breaks or not substrings:
                return None
            substrings = [line[: breaks[0]]] + substrings
            substrings = filter(len, substrings)
            substrings = map(str.strip, substrings)
            return ElectiveCell(value=list(substrings), a1=a1)

        df = df.map(lambda x: process_line(x) if isinstance(x, str) else x)

        for date, date_column in df.items():
            if not isinstance(date, datetime.date):
                warnings.warn(f"Expected date as index, got {type(date).__name__}")
                continue
            for timeslot, cell in date_column.items():
                if not (
                    isinstance(timeslot, tuple)
                    and len(timeslot) == 2
                    and all(isinstance(t, datetime.time) for t in timeslot)
                ):
                    warnings.warn(f"Expected timeslot as tuple of two datetime.time, got {timeslot!r}")
                    continue

                if isinstance(cell, ElectiveCell):
                    yield from convert_cell_to_events(cell, date, timeslot, electives, sheet_name)
